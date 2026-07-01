#!/usr/bin/env python3
"""
edwin_incident_lookup.py

Looks up Edwin AI / DEXDA incident numbers for a batch of LogicMonitor
source alert IDs (the DS/ES/LME-style identifiers shown in LM alert exports).

Auth
----
Uses the OAuth 2.2 Client Credentials flow documented by LogicMonitor for
Edwin AI API credentials (see the Edwin AI Datadog Integration doc, which
shows the same client/secret pair used against an "Access Token URL" of
the form https://<account>.dexda.ai/auth/token).

Never hardcode the client secret in this file. Set it via environment
variables instead:

    export EDWIN_CLIENT_ID="<your-client-id>"
    export EDWIN_CLIENT_SECRET="<your secret>"
    export EDWIN_ACCOUNT="<your-account>"

Usage
-----
    python3 edwin_incident_lookup.py DS0000001 DS0000002 DS0000003
    python3 edwin_incident_lookup.py --file source_ids.txt
    python3 edwin_incident_lookup.py --file source_ids.txt --excel results.xlsx

Notes / assumptions worth double-checking on first run
--------------------------------------------------------
1. Token endpoint path (/auth/token) is taken from LogicMonitor's published
   Datadog integration doc. If your tenant 401s on get_access_token(),
   the most likely fix is adjusting TOKEN_PATH below -- check the response
   body of that first call, it usually says what's wrong.
2. The query endpoint (/ui/query/records) and filter schema were captured
   live from the Edwin AI dashboard's network traffic, so that part is
   confirmed working for source-ID lookups on a production tenant.
3. Incident field names CONFIRMED via live --raw output on a production tenant:
     - cf.eventSourceId      -> the DS/ES source ID (e.g. "DS0000000")
     - alertDetails.incidentId   -> e.g. "INC0000000"
     - alertDetails.incidentUrl  -> direct ServiceNow link
     - alertDetails.workflowState -> e.g. "incident-active", "incident-resolution..."
   IMPORTANT: these are FLAT keys containing literal dots (e.g. the key
   in the JSON dict is literally "cf.eventSourceId"), not nested dicts.
   Do not try to walk them with dict.get("cf", {}).get("eventSourceId").
"""

import argparse
import csv
import json
import os
import sys
from typing import Iterable

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ACCOUNT = os.environ.get("EDWIN_ACCOUNT", "<your-account>")
CLIENT_ID = os.environ.get("EDWIN_CLIENT_ID")
CLIENT_SECRET = os.environ.get("EDWIN_CLIENT_SECRET")

BASE_URL = f"https://{ACCOUNT}.dexda.ai"
TOKEN_PATH = "/auth/token"           # per LM's Datadog integration doc
QUERY_PATH = "/ui/query/records"     # confirmed via live network capture

# Confirmed real field names from a live --raw capture on a production tenant.
# These are FLAT keys (the literal string contains dots) -- do not treat
# as nested dict paths.
SOURCE_ID_FIELD = "cf.eventSourceId"
INCIDENT_FIELDS = [
    "alertDetails.incidentId",
    "alertDetails.incidentUrl",
    "alertDetails.workflowState",
    "meta.insightKeyList",  # bare key, e.g. "correlation_<uuid>" -- NOTE this
                             # is missing the trailing "_<timestamp>" segment
                             # that the real insight record id has, so it
                             # can't be used directly with query_insight_by_key()
                             # without resolving the full id first.
]

# Confirmed via live network capture: GET /ui/query/record/insights/{id}
# returns a record with this shape. Mirrors the alertDetails.* convention.
INSIGHT_FIELDS = [
    "insightDetails.incidentId",
    "insightDetails.incidentUrl",
    "insightDetails.workflowState",
]


def get_access_token() -> str:
    if not CLIENT_ID or not CLIENT_SECRET:
        sys.exit(
            "Missing EDWIN_CLIENT_ID / EDWIN_CLIENT_SECRET environment "
            "variables. Set them before running this script -- never "
            "paste the secret directly into this file."
        )

    resp = requests.post(
        f"{BASE_URL}{TOKEN_PATH}",
        data={
            "grant_type": "client_credentials",
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
        },
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=30,
    )

    if resp.status_code != 200:
        sys.exit(
            f"Token request failed ({resp.status_code}): {resp.text}\n"
            f"Check TOKEN_PATH in this script against your tenant's actual "
            f"OAuth endpoint if this persists."
        )

    token_data = resp.json()
    access_token = token_data.get("access_token")
    if not access_token:
        sys.exit(f"Token response had no access_token field: {token_data}")
    return access_token


def query_alerts_by_source_id(
    token: str, source_ids: list[str], size: int = 200, chunk_size: int = 100
) -> dict:
    """
    Queries Edwin AI alert records filtered by cf.eventSourceId IN [...].

    Chunks large ID lists (chunk_size per request) since the IN filter's
    behavior with very large arrays is unconfirmed/untested -- chunking
    keeps each request small and reliable rather than risking a silent
    truncation or a 400 on an oversized payload. Results from all chunks
    are merged into a single combined response.

    NOTE: omits the time-window (meta.firstEventTimestamp WITHIN) clause
    seen in the dashboard UI's default query, since for a targeted ID
    lookup you generally don't want results silently dropped just because
    the alert is older than the dashboard's selected range. If your tenant
    requires that clause to be present, add it back into "expression".
    """
    combined_results: list[dict] = []

    for i in range(0, len(source_ids), chunk_size):
        chunk = source_ids[i : i + chunk_size]
        body = {
            "env": {"timezone": "UTC"},
            "source": "alerts",
            "filter": {
                "schemaName": "filterCondition",
                "schemaVersion": 4,
                "expression": {
                    "IN": [
                        {"field": "cf.eventSourceId", "type": "string"},
                        chunk,
                    ]
                },
            },
            "elasticQuery": None,
            "fields": None,
            "order": [],
            "size": size,
        }

        resp = requests.post(
            f"{BASE_URL}{QUERY_PATH}",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
            json=body,
            timeout=30,
        )

        if resp.status_code != 200:
            sys.exit(f"Query failed ({resp.status_code}) on chunk starting at index {i}: {resp.text}")

        combined_results.extend(resp.json().get("results", []))

    return {"meta": {"count": len(combined_results), "recordType": "alerts"}, "results": combined_results}


def query_alerts_by_ci_object(token: str, ci: str, obj: str, name: str | None = None) -> dict:
    """
    Fallback lookup for cases where the original source ID was a
    superseded intermediate firing (cf.eventSourceId gets overwritten
    each time the same alert condition re-fires, but the underlying
    monitored object/datapoint stays constant). Searches by
    cf.eventCI + cf.eventObject (+ optionally cf.eventName) instead,
    which are confirmed stable across an alert's full lifetime.

    This is the right fix for the "92 not found" class of problem,
    confirmed via a live alert showing the same CI/Object/Name across
    15 events spanning 9 months while cf.eventSourceId changed multiple
    times within that same window.
    """
    clauses = [
        {"EQUALS": [{"field": "cf.eventCI", "type": "string"}, ci]},
        {"EQUALS": [{"field": "cf.eventObject", "type": "string"}, obj]},
    ]
    if name:
        clauses.append({"EQUALS": [{"field": "cf.eventName", "type": "string"}, name]})

    body = {
        "env": {"timezone": "UTC"},
        "source": "alerts",
        "filter": {
            "schemaName": "filterCondition",
            "schemaVersion": 4,
            "expression": {"AND": clauses},
        },
        "elasticQuery": None,
        "fields": None,
        "order": [],
        "size": 50,
    }

    resp = requests.post(
        f"{BASE_URL}{QUERY_PATH}",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json=body,
        timeout=30,
    )

    if resp.status_code != 200:
        sys.exit(f"CI/Object query failed ({resp.status_code}): {resp.text}")

    return resp.json()


def query_alerts_by_internal_id(token: str, internal_id: str) -> dict:
    """
    EXPLORATORY: tests whether raw.sourceRecord.internalId (the stable
    LMD-style ID, e.g. "LMD48049686") is filterable, the same way
    cf.eventSourceId is. internalId stays constant across repeated
    re-firings of the same underlying alert condition over time, even
    though cf.eventSourceId gets overwritten with a new DS/ES number each
    time it re-fires -- confirmed via a live raw record showing 15 events
    sharing one internalId but multiple different source IDs across them.

    UNCONFIRMED: whether the filter engine can reach into the JSON string
    stored at raw.sourceRecord to filter on internalId, since unlike
    cf.eventSourceId this isn't a flat top-level key -- it's a field
    inside a JSON-encoded string value. This may simply not work.
    """
    body = {
        "env": {"timezone": "UTC"},
        "source": "alerts",
        "filter": {
            "schemaName": "filterCondition",
            "schemaVersion": 4,
            "expression": {
                "EQUALS": [
                    {"field": "raw.sourceRecord.internalId", "type": "string"},
                    internal_id,
                ]
            },
        },
        "elasticQuery": None,
        "fields": None,
        "order": [],
        "size": 50,
    }

    resp = requests.post(
        f"{BASE_URL}{QUERY_PATH}",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json=body,
        timeout=30,
    )

    if resp.status_code != 200:
        sys.exit(
            f"internalId query failed ({resp.status_code}): {resp.text}\n"
            f"This likely confirms raw.sourceRecord.internalId isn't a "
            f"filterable field via this path -- the JSON-string nesting "
            f"probably isn't reachable by the filter engine."
        )

    return resp.json()


def query_insight_by_key(token: str, insight_key: str) -> dict:
    """
    Looks up insight record(s) by key, using the bulk query endpoint
    (same /ui/query/records endpoint the alert lookup uses) filtered on
    meta.rowKey. This accepts the BARE key format as seen in an alert's
    meta.insightKeyList field (e.g. "correlation_<uuid>", no trailing
    timestamp) -- confirmed via live capture that an insight record's
    meta.rowKey field matches this bare format exactly, even though the
    record's own _id has an extra "_<timestamp>" suffix.

    (There's also a direct GET /ui/query/record/insights/{full_id}
    single-record endpoint, confirmed working when you already have the
    FULL id with timestamp suffix -- e.g. from a UI URL. This function
    uses the bulk-query path instead since that's what works from the
    bare key alone, which is all alert records actually expose.)
    """
    body = {
        "env": {"timezone": "UTC"},
        "source": "insights",
        "filter": {
            "schemaName": "filterCondition",
            "schemaVersion": 4,
            "expression": {
                "EQUALS": [
                    {"field": "meta.rowKey", "type": "string"},
                    insight_key,
                ]
            },
        },
        "elasticQuery": None,
        "fields": None,
        "order": [],
        "size": 10,
    }

    resp = requests.post(
        f"{BASE_URL}{QUERY_PATH}",
        headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
        },
        json=body,
        timeout=30,
    )

    if resp.status_code != 200:
        sys.exit(f"Insight query failed ({resp.status_code}): {resp.text}")

    return resp.json()


def get_insight_by_full_id(token: str, full_insight_id: str) -> dict:
    """
    Direct single-record lookup, confirmed via live network capture of
    Edwin's Explore page: GET /ui/query/record/insights/{full_id}

    Requires the FULL id including the trailing timestamp suffix (e.g.
    "correlation_8c4f77f5-c1c5-429c-b47a-b88bcd478660_1781746816456"),
    not the bare value from meta.insightKeyList. Use
    query_insight_by_key() instead when you only have the bare key.
    """
    resp = requests.get(
        f"{BASE_URL}/ui/query/record/insights/{full_insight_id}",
        headers={"Authorization": f"Bearer {token}"},
        timeout=30,
    )

    if resp.status_code != 200:
        sys.exit(f"Insight lookup failed ({resp.status_code}): {resp.text}")

    return resp.json()


def summarize(results: dict) -> list[dict]:
    rows = []
    for record in results.get("results", []):
        source_id = record.get(SOURCE_ID_FIELD)
        found = {field: record[field] for field in INCIDENT_FIELDS if record.get(field)}
        rows.append(
            {
                "source_id": source_id,
                "incident_fields_found": found,
            }
        )
    return rows


def group_by_source_id(rows: list[dict]) -> list[dict]:
    """
    Collapses multiple rows sharing the same Source ID into a single row.
    Some source IDs return more than one alert record (e.g. recurring
    alerts on the same DS/ES number at different times), sometimes with
    different linked incidents. Rather than silently pick one, this joins
    all distinct non-empty values per field with " | " so nothing is lost
    -- if you see multiple incident numbers joined together, that's a
    genuine sign the same source ID had different incident outcomes
    across occurrences and is worth a manual look.
    """
    grouped: dict[str, dict] = {}
    order: list[str] = []

    for row in rows:
        sid = row["source_id"]
        if sid not in grouped:
            grouped[sid] = {field: set() for field in INCIDENT_FIELDS}
            order.append(sid)
        for field in INCIDENT_FIELDS:
            val = row["incident_fields_found"].get(field)
            if val:
                grouped[sid][field].add(str(val))

    result = []
    for sid in order:
        found = {
            field: " | ".join(sorted(vals))
            for field, vals in grouped[sid].items()
            if vals
        }
        result.append({"source_id": sid, "incident_fields_found": found})
    return result


def _resolve_one_ci_object(token: str, sid: str, ci: str, obj: str) -> dict:
    """Resolves a single (Source ID, CI, Object) triple via the CI+Object
    fallback search and returns a row in the standard shape."""
    result = query_alerts_by_ci_object(token, ci, obj)
    records = result.get("results", [])
    # Prefer the record actually showing an incidentId, since a CI+Object
    # pair can match multiple historical alert records (a CI+Object pair
    # can return multiple results; prefer the one with an incidentId).
    best = next(
        (r for r in records if r.get("alertDetails.incidentId")), None
    ) or (records[0] if records else None)

    found = {}
    if best:
        found = {field: best[field] for field in INCIDENT_FIELDS if best.get(field)}
    return {"source_id": sid, "incident_fields_found": found}


def resolve_ci_object_batch(token: str, path: str) -> list[dict]:
    """
    Reads a CSV with columns "Source ID,CI,Object" (header required, in
    that order) and resolves each row via the CI+Object fallback search
    -- for the "92 not found" class of problem, where the original
    cf.eventSourceId was a superseded intermediate firing.

    Fill this file in by hand: for each ID in the "Needs Manual Check"
    sheet, open its LogicMonitor search link, read off the device/CI name
    and the datasource instance (Object), and add a row here. Example:

        Source ID,CI,Object
        DS0000000,<device-name>,<datasource-instance>

    Returns rows in the same shape as summarize()/group_by_source_id(),
    so they can be merged into the normal output flow.
    """
    rows = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f)
        for line in reader:
            sid = line.get("Source ID", "").strip()
            ci = line.get("CI", "").strip()
            obj = line.get("Object", "").strip()
            if not sid or not ci or not obj:
                continue
            rows.append(_resolve_one_ci_object(token, sid, ci, obj))

    return rows


def resolve_from_active_alerts_csv(token: str, path: str, target_ids: list[str]) -> list[dict]:
    """
    Reads an export from lm_active_alerts.py (columns: id, monitorObjectName,
    instanceName, ...) and resolves each ID in target_ids via the CI+Object
    fallback search -- no manual lookup needed, since this export already
    has the CI (monitorObjectName) and Object (instanceName) for every
    source ID it covers.

    Only resolves IDs that are both in target_ids AND present in this CSV;
    IDs missing from the CSV simply aren't returned (caller's existing
    missing-IDs handling covers those same as before).
    """
    target_set = set(target_ids)
    rows = []
    with open(path, newline="") as f:
        reader = csv.DictReader(f)
        for line in reader:
            sid = line.get("id", "").strip()
            if sid not in target_set:
                continue
            ci = line.get("monitorObjectName", "").strip()
            obj = line.get("instanceName", "").strip()
            if not ci or not obj:
                continue
            rows.append(_resolve_one_ci_object(token, sid, ci, obj))

    return rows


def load_snow_states(path: str) -> tuple[dict[str, str], dict[str, tuple[str, str]]]:
    """
    Reads a ServiceNow incident export (.xlsx) and returns two dicts:

    1. snow_states: Incident Number -> State
       e.g. {"INC0000000": "In Progress"}

    2. snow_src_id_lookup: LM Source ID -> (Incident Number, State)
       e.g. {"ES0000000": ("INC0000000", "New")}
       Built by extracting DS/ES source IDs from the Description(description)
       column, which contains LogicMonitor alert URLs in the format:
       "LogicMonitor Alert: https://.../alerts/ES0000000"
       This gives a direct, reliable source-ID match for incidents where
       Edwin has no link -- confirmed against real SNOW export data showing
       many unique source IDs present in this field.

    Expects columns "Number" and "State" (required) and
    "Description(description)" (optional, used for src_id lookup if present).
    """
    import re
    src_id_pattern = re.compile(r'/alerts/([A-Z]{2}\d+)', re.IGNORECASE)

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {str(val).strip(): idx for idx, val in enumerate(header_row) if val}

    if "Number" not in header_map or "State" not in header_map:
        sys.exit(
            f"Could not find 'Number' and 'State' columns in {path}.\n"
            f"Columns found: {list(header_map.keys())}"
        )

    number_idx = header_map["Number"]
    state_idx = header_map["State"]
    desc_idx = header_map.get("Description(description)")

    snow_states: dict[str, str] = {}
    snow_src_id_lookup: dict[str, tuple[str, str]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        number = row[number_idx]
        state = row[state_idx]
        if not number:
            continue
        num_str = str(number).strip()
        state_str = str(state).strip() if state else ""
        snow_states[num_str] = state_str

        if desc_idx is not None:
            desc = str(row[desc_idx] or "")
            for src_id in src_id_pattern.findall(desc):
                sid_upper = src_id.upper()
                if sid_upper not in snow_src_id_lookup:
                    snow_src_id_lookup[sid_upper] = (num_str, state_str)

    return snow_states, snow_src_id_lookup


def build_snow_ci_lookup(snow_path: str) -> list[dict]:
    """
    Reads a ServiceNow incident export (.xlsx) and builds a list of
    incident records keyed by CI and Short description text, for use
    as a last-resort fallback when Edwin has no incident link at all.

    Match logic (confirmed against real data):
    - Configuration item (lowercased) must match alert's monitorObjectName
    - Short description must contain "{dataPointName}.{resourceTemplateName}"
      (lowercased, case-insensitive) -- this is the format LogicMonitor
      writes to SNOW short descriptions:
      e.g. "... <device-name> : <DataPoint>.<DataSource>.dataSourceAlert"

    Returns a list of dicts with keys:
      ci_lower, match_key_lower, Number, State, ShortDescription
    """
    wb = load_workbook(snow_path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {str(val).strip(): idx for idx, val in enumerate(header_row) if val}

    needed = ["Number", "State", "Configuration item", "Short description"]
    missing = [c for c in needed if c not in header_map]
    if missing:
        sys.exit(
            f"Could not find columns {missing} in {snow_path}.\n"
            f"Columns found: {list(header_map.keys())}"
        )

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        number = row[header_map["Number"]]
        state = row[header_map["State"]]
        ci = row[header_map["Configuration item"]]
        desc = row[header_map["Short description"]]
        if not number or not ci or not desc:
            continue
        records.append({
            "ci_lower": str(ci).strip().lower(),
            "desc_lower": str(desc).strip().lower(),
            "Number": str(number).strip(),
            "State": str(state).strip() if state else "",
            "ShortDescription": str(desc).strip(),
        })

    return records


def snow_fallback_lookup(
    snow_records: list[dict],
    monitor_object: str,
    resource_template: str,
    datapoint: str,
) -> list[dict]:
    """
    Searches snow_records for incidents matching this alert's
    CI + dataPointName.resourceTemplateName combination.
    Returns all matching incident records (caller handles 0/1/many).
    """
    ci_lower = monitor_object.strip().lower()
    match_key = f"{datapoint.strip()}.{resource_template.strip()}".lower()
    return [
        r for r in snow_records
        if r["ci_lower"] == ci_lower and match_key in r["desc_lower"]
    ]


def write_excel(rows: list[dict], missing_ids: list[str], path: str, snow_states: dict[str, str] | None = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Incident Lookup"

    headers = ["Source ID", "Incident", "Edwin State", "ServiceNow State", "Link", "Insight Key", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="44546A")

    for row in rows:
        fields = row["incident_fields_found"]
        direct_incident = fields.get("alertDetails.incidentId", "")
        resolved_incident = fields.get("resolved_via_insight.incidentId", "")

        if direct_incident:
            incident_id = direct_incident
            edwin_state = fields.get("alertDetails.workflowState", "")
            link = fields.get("alertDetails.incidentUrl", "")
            status = "Multiple incidents seen" if " | " in direct_incident else "Linked"
        elif resolved_incident:
            incident_id = resolved_incident
            edwin_state = fields.get("resolved_via_insight.workflowState", "")
            link = fields.get("resolved_via_insight.incidentUrl", "")
            status = "Linked via insight"
        elif fields.get("snow_fallback.incidentId"):
            incident_id = fields["snow_fallback.incidentId"]
            edwin_state = ""
            link = ""
            status = "Linked via SNOW match" if " | " not in incident_id else "Multiple via SNOW match"
        else:
            incident_id = ""
            edwin_state = ""
            link = ""
            status = "No incident linked"

        if snow_states and incident_id:
            # For SNOW-fallback rows, state is already embedded directly
            if fields.get("snow_fallback.incidentId") and not direct_incident and not resolved_incident:
                snow_state = fields.get("snow_fallback.state", "")
            else:
                snow_state = " | ".join(
                    sorted(set(
                        snow_states.get(inc.strip(), "Not in SNOW export")
                        for inc in incident_id.split(" | ")
                    ))
                )
        else:
            snow_state = ""

        ws.append(
            [
                row["source_id"],
                incident_id,
                edwin_state,
                snow_state,
                link,
                fields.get("meta.insightKeyList", ""),
                status,
            ]
        )

    for sid in missing_ids:
        ws.append([sid, "", "", "", "", "", "No alert record found"])

    widths = [16, 16, 18, 18, 60, 30, 22]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    if missing_ids:
        ws2 = wb.create_sheet("Needs Manual Check")
        ws2.append(["Source ID", "LogicMonitor Search Link", "Note"])
        for col in range(1, 4):
            cell = ws2.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="44546A")

        note = (
            "Not found via cf.eventSourceId -- likely a superseded "
            "intermediate firing of a recurring alert. To resolve: open "
            "the link, read off the CI (device name) and Object "
            "(datasource instance) shown there, add a row for this ID to "
            "a CSV with columns 'Source ID,CI,Object', then re-run this "
            "script with --ci-object-file pointing at that CSV -- it will "
            "batch-resolve the real incident number automatically."
        )
        for sid in missing_ids:
            ws2.append([sid, f"https://{ACCOUNT}.logicmonitor.com/santaba/uiv4/alerts/{sid}", note])

        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 60
        ws2.column_dimensions["C"].width = 80

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(
        description="Look up Edwin AI incidents by LM source alert ID (DS/ES numbers)."
    )
    parser.add_argument("source_ids", nargs="*", help="DS/ES source IDs, e.g. DS0000000")
    parser.add_argument("--file", help="Path to a text file, one source ID per line")
    parser.add_argument(
        "--raw", action="store_true", help="Print full raw API response instead of summary"
    )
    parser.add_argument(
        "--excel", help="Write results to an .xlsx file at this path, in addition to console output"
    )
    parser.add_argument(
        "--test-insight-key",
        help="Look up insight record(s) by BARE key (the value seen in an "
        "alert's meta.insightKeyList field), e.g. "
        "'correlation_a13d0991-7573-413c-947b-7df2de98b3b8'. Prints raw JSON.",
    )
    parser.add_argument(
        "--resolve-insights",
        action="store_true",
        help="For alerts with no direct incident link but a populated Insight "
        "Key, automatically look up the linked insight and pull its "
        "incident number from there. Adds one extra API call per such "
        "alert, so off by default.",
    )
    parser.add_argument(
        "--snow-export",
        help="Path to a ServiceNow incident export (.xlsx) with 'Number' and "
        "'State' columns. When provided, adds a 'ServiceNow State' column "
        "showing each incident's actual current state (New/In Progress/"
        "On Hold/etc.) instead of just Edwin's own workflow state. "
        "Incidents not found in this export are flagged 'Not in SNOW "
        "export' (often means closed/resolved since the export was taken).",
    )
    parser.add_argument(
        "--test-internal-id",
        help="EXPLORATORY: test whether an LMD-style internal ID (e.g. "
        "'LMD48049686', found in an alert's description text) is "
        "filterable as raw.sourceRecord.internalId. Prints raw JSON.",
    )
    parser.add_argument(
        "--test-ci-object",
        nargs=2,
        metavar=("CI", "OBJECT"),
        help="Test the CI+Object fallback search, e.g. "
        "--test-ci-object <device-name> '<datasource-instance>' "
        "Prints raw JSON.",
    )
    parser.add_argument(
        "--ci-object-file",
        help="CSV with columns 'Source ID,CI,Object' for resolving alerts "
        "whose original source ID was a superseded intermediate firing "
        "(the 'Needs Manual Check' cases). Fill in CI/Object by hand from "
        "each ID's LogicMonitor search link, then run with this flag -- "
        "results merge into the normal --excel/console output as if they "
        "were resolved directly.",
    )
    parser.add_argument(
        "--active-alerts-csv",
        help="Export from lm_active_alerts.py (columns: id, "
        "monitorObjectName, instanceName, ...). When provided, any source "
        "ID not found via the normal lookup is automatically resolved "
        "using this file's CI/Object data instead -- no manual lookup "
        "needed, since this export already has everything required.",
    )
    args = parser.parse_args()

    if args.test_ci_object:
        token = get_access_token()
        ci, obj = args.test_ci_object
        result = query_alerts_by_ci_object(token, ci, obj)
        print(json.dumps(result, indent=2))
        return

    if args.test_internal_id:
        token = get_access_token()
        result = query_alerts_by_internal_id(token, args.test_internal_id)
        print(json.dumps(result, indent=2))
        return

    if args.test_insight_key:
        token = get_access_token()
        result = query_insight_by_key(token, args.test_insight_key)
        print(json.dumps(result, indent=2))
        return

    source_ids: list[str] = list(args.source_ids)
    if args.file:
        with open(args.file) as f:
            source_ids.extend(line.strip() for line in f if line.strip())

    # active_alerts.csv can serve double duty: supplying source IDs for
    # the primary Edwin lookup AND providing CI/Object metadata for the
    # SNOW fallback -- no need for a separate source_ids.txt file.
    if args.active_alerts_csv and not source_ids and not args.ci_object_file:
        with open(args.active_alerts_csv, newline="") as f:
            reader = csv.DictReader(f)
            source_ids.extend(
                line["id"].strip() for line in reader
                if line.get("id", "").strip()
            )

    if not source_ids and not args.ci_object_file:
        parser.error("Provide source IDs as arguments, via --file, via --active-alerts-csv, or via --ci-object-file")

    token = get_access_token()

    rows: list[dict] = []
    if source_ids:
        results = query_alerts_by_source_id(token, source_ids)

        if args.raw:
            print(json.dumps(results, indent=2))
            return

        rows = summarize(results)
        rows = group_by_source_id(rows)

    if args.ci_object_file:
        ci_object_rows = resolve_ci_object_batch(token, args.ci_object_file)
        rows.extend(ci_object_rows)
        # IDs resolved this way came from the manual-check file, not the
        # main source_ids list -- include them so they're not also
        # reported as "missing" below.
        source_ids.extend(row["source_id"] for row in ci_object_rows)

    if args.resolve_insights:
        for row in rows:
            fields = row["incident_fields_found"]
            insight_key = fields.get("meta.insightKeyList")
            if insight_key and not fields.get("alertDetails.incidentId"):
                # insight_key may itself be multiple " | "-joined keys after
                # grouping; resolve each and merge in whatever's found.
                for key in insight_key.split(" | "):
                    insight_result = query_insight_by_key(token, key.strip())
                    insight_records = insight_result.get("results", [])
                    if insight_records:
                        rec = insight_records[0]
                        if rec.get("insightDetails.incidentId"):
                            fields["resolved_via_insight.incidentId"] = rec["insightDetails.incidentId"]
                        if rec.get("insightDetails.incidentUrl"):
                            fields["resolved_via_insight.incidentUrl"] = rec["insightDetails.incidentUrl"]
                        if rec.get("insightDetails.workflowState"):
                            fields["resolved_via_insight.workflowState"] = rec["insightDetails.workflowState"]

    found_ids = {row["source_id"] for row in rows}
    missing_ids = [sid for sid in source_ids if sid not in found_ids]

    if args.active_alerts_csv and missing_ids:
        auto_resolved = resolve_from_active_alerts_csv(token, args.active_alerts_csv, missing_ids)
        rows.extend(auto_resolved)
        found_ids = {row["source_id"] for row in rows}
        missing_ids = [sid for sid in source_ids if sid not in found_ids]

    snow_states, snow_src_id_lookup = load_snow_states(args.snow_export) if args.snow_export else ({}, {})

    # NEW: SNOW Description direct source-ID match.
    # The SNOW export's Description(description) column contains LogicMonitor
    # alert URLs (e.g. ".../alerts/ES0000000") -- confirmed many unique source
    # IDs present across real export data. This is more reliable than the
    # CI+datasource match since it's an exact ID lookup, no fuzzy matching.
    if snow_src_id_lookup:
        for row in rows:
            fields = row["incident_fields_found"]
            already_linked = (
                fields.get("alertDetails.incidentId") or
                fields.get("resolved_via_insight.incidentId")
            )
            if already_linked:
                continue
            match = snow_src_id_lookup.get(row["source_id"].upper())
            if match:
                fields["snow_fallback.incidentId"] = match[0]
                fields["snow_fallback.state"] = match[1]

    # SNOW CI+datasource fallback: for rows that Edwin found (alert record
    # exists) but still has no incident link after the direct ID match,
    # try matching against the SNOW export using CI + dataPointName.
    # resourceTemplateName from active_alerts.csv.
    if snow_states and args.active_alerts_csv:
        snow_records = build_snow_ci_lookup(args.snow_export)
        alert_meta: dict[str, dict] = {}
        with open(args.active_alerts_csv, newline="") as f:
            reader = csv.DictReader(f)
            for line in reader:
                sid = line.get("id", "").strip()
                if sid:
                    alert_meta[sid] = {
                        "monitorObjectName": line.get("monitorObjectName", "").strip(),
                        "resourceTemplateName": line.get("resourceTemplateName", "").strip(),
                        "dataPointName": line.get("dataPointName", "").strip(),
                    }

        for row in rows:
            fields = row["incident_fields_found"]
            already_linked = (
                fields.get("alertDetails.incidentId") or
                fields.get("resolved_via_insight.incidentId") or
                fields.get("snow_fallback.incidentId")
            )
            if already_linked:
                continue
            meta = alert_meta.get(row["source_id"])
            if not meta or not meta["monitorObjectName"]:
                continue
            matches = snow_fallback_lookup(
                snow_records,
                meta["monitorObjectName"],
                meta["resourceTemplateName"],
                meta["dataPointName"],
            )
            if matches:
                incident_ids = " | ".join(dict.fromkeys(m["Number"] for m in matches))
                states = " | ".join(dict.fromkeys(m["State"] for m in matches))
                fields["snow_fallback.incidentId"] = incident_ids
                fields["snow_fallback.state"] = states

    for row in rows:
        print(f"\nSource ID: {row['source_id']}")
        incident_id = None
        if row["incident_fields_found"].get("alertDetails.incidentId"):
            incident_id = row['incident_fields_found']['alertDetails.incidentId']
            print(f"  Incident: {incident_id}")
            if row["incident_fields_found"].get("alertDetails.workflowState"):
                print(f"  Edwin State: {row['incident_fields_found']['alertDetails.workflowState']}")
            if row["incident_fields_found"].get("alertDetails.incidentUrl"):
                print(f"  Link: {row['incident_fields_found']['alertDetails.incidentUrl']}")
        elif row["incident_fields_found"].get("resolved_via_insight.incidentId"):
            incident_id = row['incident_fields_found']['resolved_via_insight.incidentId']
            print(f"  Incident: {incident_id} (via correlated insight)")
            if row["incident_fields_found"].get("resolved_via_insight.workflowState"):
                print(f"  Edwin State: {row['incident_fields_found']['resolved_via_insight.workflowState']}")
            if row["incident_fields_found"].get("resolved_via_insight.incidentUrl"):
                print(f"  Link: {row['incident_fields_found']['resolved_via_insight.incidentUrl']}")
        else:
            fallback_inc = row["incident_fields_found"].get("snow_fallback.incidentId")
            if fallback_inc:
                incident_id = fallback_inc
                print(f"  Incident: {fallback_inc} (via SNOW CI/datasource match)")
                fallback_state = row["incident_fields_found"].get("snow_fallback.state", "")
                if fallback_state:
                    print(f"  ServiceNow State: {fallback_state}")
            else:
                print("  No incident linked to this alert.")
        if snow_states and incident_id:
            snow_vals = sorted(set(
                snow_states.get(inc.strip(), "Not in SNOW export")
                for inc in incident_id.split(" | ")
            ))
            print(f"  ServiceNow State: {' | '.join(snow_vals)}")
        if row["incident_fields_found"].get("meta.insightKeyList"):
            print(f"  Insight key: {row['incident_fields_found']['meta.insightKeyList']}")

    if missing_ids:
        print(f"\nNo alert record found at all for: {', '.join(missing_ids)}")
        print("(likely resolved/cleared and outside Edwin's currently-queryable window,")
        print(" or the ID doesn't exist as an alert in this tenant)")

    if args.excel:
        write_excel(rows, missing_ids, args.excel, snow_states or None)
        print(f"\nExcel file written to: {args.excel}")


if __name__ == "__main__":
    main()